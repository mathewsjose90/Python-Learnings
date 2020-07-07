import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import os


def process_workbook(filename):
    pass
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    '''
    cell = sheet['a1']
    cell1 = sheet.cell(1, 1)
    '''

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        correct_value = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = correct_value

    # For drawing chart
    value_ref = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(value_ref)
    sheet.add_chart(chart, 'e2')
    old_name, ext = os.path.splitext(filename)
    new_filename = old_name + '_v2' + ext
    wb.save(new_filename)


process_workbook('transactions_ex.xlsx')
